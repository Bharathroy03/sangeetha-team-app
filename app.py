import os
import json
from datetime import datetime as real_datetime, timezone, timedelta

class datetime(real_datetime):
    @classmethod
    def now(cls, tz=None):
        # Always return Asia/Kolkata time (UTC+5:30) as naive datetime
        utc_now = real_datetime.now(timezone.utc)
        kolkata_tz = timezone(timedelta(hours=5, minutes=30))
        return utc_now.astimezone(kolkata_tz).replace(tzinfo=None)

from functools import wraps
from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
import io
import csv
from flask import send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
import logging
from logging.handlers import RotatingFileHandler
import traceback

app = Flask(__name__)
# In production, this secret key should be loaded from environment variables
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'sangeetha_super_secret_session_key')

# Configure Logging
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
log_formatter = logging.Formatter('[%(asctime)s] %(levelname)s in %(module)s: %(message)s')

try:
    os.makedirs(DATA_DIR, exist_ok=True)
    log_file = os.path.join(DATA_DIR, 'app.log')
    file_handler = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=3, encoding='utf-8')
except (OSError, PermissionError):
    try:
        # Fallback to /tmp on serverless environments like Vercel
        tmp_log_file = os.path.join('/tmp', 'app.log')
        file_handler = RotatingFileHandler(tmp_log_file, maxBytes=5*1024*1024, backupCount=3, encoding='utf-8')
    except (OSError, PermissionError):
        # Ultimate fallback to StreamHandler (stdout)
        import sys
        file_handler = logging.StreamHandler(sys.stdout)

file_handler.setFormatter(log_formatter)
file_handler.setLevel(logging.INFO)
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO)
app.logger.info("Sangeetha Mobiles App Started")

def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def log_application_error(error_message, exception=None, status_code=500):
    """Log structured application errors with request context and stack trace."""
    user = session.get('username') or 'Anonymous'
    emp_id = session.get('employee_id') or 'N/A'
    role = session.get('role') or 'N/A'
    route = request.path
    method = request.method
    
    stack_trace = ""
    if exception:
        stack_trace = "".join(traceback.format_exception(type(exception), exception, exception.__traceback__))
        
    log_msg = (
        f"Route: {method} {route} | "
        f"User: {user} ({emp_id}) | "
        f"Role: {role} | "
        f"Status: {status_code} | "
        f"Error: {error_message}"
    )
    if stack_trace:
        log_msg += f"\nStack Trace:\n{stack_trace}"
        
    app.logger.error(log_msg)


# Path to our customers JSON database
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
CUSTOMERS_FILE = os.path.join(DATA_DIR, 'customers.json')
PAYMENTS_FILE = os.path.join(DATA_DIR, 'payments.json')
USERS_FILE = os.path.join(DATA_DIR, 'users.json')
EDIT_REQUESTS_FILE = os.path.join(DATA_DIR, 'customer_edit_requests.json')
CRM_WALKIN_CUSTOMERS_FILE = os.path.join(DATA_DIR, 'crm_walkin_customers.json')

WHITELIST_FILES = {
    'customers.json': CUSTOMERS_FILE,
    'crm_walkin_customers.json': CRM_WALKIN_CUSTOMERS_FILE,
    'customer_edit_requests.json': EDIT_REQUESTS_FILE,
    'users.json': USERS_FILE,
    'payments.json': PAYMENTS_FILE
}

# Supabase HTTP DB client imports
from database import (
    db_load_users, db_save_users, db_load_customers, db_save_customers,
    db_get_customers_created_today, db_get_finance_customers,
    db_load_crm_walkin, db_save_crm_walkin, db_load_payments, db_save_payments,
    db_delete_payment_history, db_load_edit_requests, db_save_edit_requests,
    db_load_audit_log, db_add_audit_log, db_count_audit_logs, db_delete_customer_record
)


from permissions import ROLE_PERMISSIONS

def load_users():
    """Load user list from Supabase."""
    users = db_load_users()
    if not users:
        # Check/seed defaults
        defaults = [
            ("USR-1001", "Bharath Kumar R", "22913", "Bharathroy@03", "super_admin"),
            ("USR-1002", "Admin", "6909", "Sang@1974", "admin"),
            ("USR-1003", "Sangeetha", "SMPL", "Sang@1974", "store_employee")
        ]
        users_to_save = []
        for uid, name, emp_id, pwd, role in defaults:
            users_to_save.append({
                "user_id": uid,
                "username": name,
                "employee_id": emp_id,
                "password_hash": generate_password_hash(pwd),
                "role": role,
                "status": "active",
                "created_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "updated_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
        db_save_users(users_to_save)
        users = db_load_users()
    return users

def save_users(users_list):
    """Save user list to Supabase."""
    db_save_users(users_list)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'employee_id' not in session or 'role' not in session:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({"success": False, "message": "Unauthorized"}), 401
            return redirect(url_for('login'))
        
        # Verify user status
        users = load_users()
        user = next((u for u in users if u.get('employee_id') == session.get('employee_id')), None)
        if not user or user.get('status') != 'active':
            session.clear()
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({"success": False, "message": "User inactive or unauthorized"}), 403
            return render_template('403.html', message="Your account is inactive."), 403
            
        return f(*args, **kwargs)
    return decorated_function

def permission_required(permission_name):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'role' not in session or 'employee_id' not in session:
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify({"success": False, "message": "Unauthorized"}), 401
                return redirect(url_for('login'))
                
            role = session['role']
            users = load_users()
            user = next((u for u in users if u.get('employee_id') == session.get('employee_id')), None)
            if not user or user.get('status') != 'active':
                session.clear()
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify({"success": False, "message": "User inactive or unauthorized"}), 403
                return render_template('403.html', message="Your account is inactive."), 403
                
            allowed_permissions = ROLE_PERMISSIONS.get(role, [])
            if permission_name not in allowed_permissions:
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify({"success": False, "message": "Forbidden: You do not have permission to access this resource."}), 403
                return render_template('403.html', message="You do not have permission to access this page."), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def load_customers():
    """Load customer list from Supabase."""
    return db_load_customers()

def save_customers(customers_list):
    """Save customer list to Supabase."""
    db_save_customers(customers_list)

def load_crm_customers():
    """Load CRM walkin customer list from Supabase."""
    return db_load_crm_walkin()

def save_crm_customers(crm_list):
    """Save CRM walkin customer list to Supabase."""
    db_save_crm_walkin(crm_list)


@app.before_request
def make_session_permanent():
    session.permanent = True

# --- VIEWS ---

@app.route('/')
def home():
    if 'employee_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login')
def login():
    if 'employee_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html', username=session['username'], role=session['role'])

@app.route('/customer-entry')
@login_required
def customer_entry():
    return render_template('customer_entry_page.html', username=session['username'], role=session['role'])

@app.route('/customer-history')
@login_required
@permission_required('customer_history_view')
def customer_history():
    return redirect(url_for('customer_records_view', tab='total'))

@app.route('/customer-records')
@login_required
@permission_required('customer_history_view')
def customer_records_view():
    return render_template('customer_records.html', username=session['username'], role=session['role'])

@app.route('/crm-data')
@login_required
@permission_required('crm_view')
def crm_data_view():
    return render_template('crm_data.html', username=session['username'], role=session['role'])

# --- API ENDPOINTS ---

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json() or {}
    employee_id = data.get('employee_id', '').strip()
    password = data.get('password', '').strip()

    if not employee_id or not password:
        return jsonify({"success": False, "message": "Employee ID and password are required."}), 400

    users = load_users()
    user = next((u for u in users if u.get('employee_id') == employee_id), None)

    if not user:
        return jsonify({"success": False, "message": "Invalid Employee ID or password."}), 401

    if user.get('status') != 'active':
        return jsonify({"success": False, "message": "Your account has been deactivated. Please contact your administrator."}), 403

    if check_password_hash(user.get('password_hash', ''), password):
        session['employee_id'] = user['employee_id']
        session['username'] = user['username']
        session['role'] = user['role']
        return jsonify({"success": True, "message": "Login successful!"})
    else:
        return jsonify({"success": False, "message": "Invalid Employee ID or password."}), 401

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({"success": True, "message": "Logged out successfully!"})

@app.route('/api/session', methods=['GET'])
@login_required
def api_session():
    users = load_users()
    user = next((u for u in users if u.get('employee_id') == session.get('employee_id')), None)
    if not user:
        return jsonify({"success": False, "message": "User not found."}), 404
    return jsonify({
        "success": True,
        "data": {
            "employee_id": user['employee_id'],
            "username": user['username'],
            "role": user['role'],
            "user_id": user['user_id'],
            "job_title": user.get('job_title', ''),
            "status": user.get('status', 'active')
        }
    })

@app.route('/api/decode-barcode', methods=['POST'])
@login_required
def api_decode_barcode():
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "No file uploaded"}), 400
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({"success": False, "message": "No file selected"}), 400
    
    try:
        from PIL import Image, ImageEnhance, UnidentifiedImageError
        from pyzbar.pyzbar import decode
        
        # Load image
        try:
            img = Image.open(file.stream)
        except UnidentifiedImageError:
            return jsonify({"success": False, "message": "Invalid or empty image file uploaded. Please upload a valid image."}), 400
        
        # Decode attempt 1: Original image
        decoded_objs = decode(img)
        
        # Decode attempt 2: Grayscale image
        if not decoded_objs:
            gray_img = img.convert('L')
            decoded_objs = decode(gray_img)
            
            # Decode attempt 3: High contrast grayscale
            if not decoded_objs:
                enhancer = ImageEnhance.Contrast(gray_img)
                enhanced_img = enhancer.enhance(2.0)
                decoded_objs = decode(enhanced_img)
        
        if decoded_objs:
            # Get the first successfully decoded barcode value
            barcode_val = decoded_objs[0].data.decode('utf-8').strip()
            return jsonify({"success": True, "barcode": barcode_val})
        else:
            return jsonify({"success": False, "message": "No barcode detected in the uploaded image. Please ensure it is clear and aligned."}), 422
            
    except Exception as e:
        app.logger.error(f"Barcode decoding API failed: {e}\n{traceback.format_exc()}")
        return jsonify({"success": False, "message": f"Decoding failed: {str(e)}"}), 500

@app.route('/api/customers/today', methods=['GET'])
@login_required
@permission_required('customer_history_view')
def api_get_customers_today():
    today_str = datetime.now().strftime('%Y-%m-%d')
    customers = db_get_customers_created_today(today_str)
    return jsonify({"success": True, "data": customers})

@app.route('/api/customers/finance', methods=['GET'])
@login_required
@permission_required('customer_history_view')
def api_get_customers_finance():
    customers = db_get_finance_customers()
    return jsonify({"success": True, "data": customers})

@app.route('/api/customers/<customer_id>', methods=['GET'])
@login_required
@permission_required('customer_history_view')
def api_get_customer_detail(customer_id):
    customers = load_customers()
    customer = next((c for c in customers if c.get('customer_id') == customer_id), None)
    if not customer:
        return jsonify({"success": False, "message": "Customer not found."}), 404
    return jsonify({"success": True, "data": customer})

from database import (
    SupabaseError, SupabaseConnectionError, SupabaseQueryError,
    SupabaseRlsError, SupabaseOperationError
)

@app.route('/403')
def route_403():
    return render_template('403.html', message="Forbidden: Access is denied."), 403

@app.errorhandler(SupabaseError)
def handle_supabase_error(error):
    status = 500
    if isinstance(error, SupabaseRlsError):
        status = 403
        msg = "Database permission denied. Contact admin."
    elif isinstance(error, SupabaseQueryError):
        status = 404
        msg = "Database query failed. Resource not found."
    elif isinstance(error, SupabaseConnectionError):
        status = 503
        msg = "Database connection failed. Contact admin."
    else:
        status = 500
        msg = "Database operation failed. Contact admin."
        
    log_application_error(f"SupabaseError ({type(error).__name__}): {error.message}", exception=error, status_code=status)
    
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({
            "success": False,
            "message": msg,
            "error_code": type(error).__name__,
            "details": getattr(error, 'details', None) or str(error)
        }), status
    return render_template('500.html', message=msg), status

@app.errorhandler(400)
def bad_request_error(error):
    msg = "Bad request. Please verify request parameters."
    log_application_error(msg, exception=error, status_code=400)
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({
            "success": False,
            "message": msg,
            "error_code": "BAD_REQUEST",
            "details": str(error)
        }), 400
    return render_template('403.html', message=msg), 400

@app.errorhandler(401)
def unauthorized_error(error):
    msg = "Session expired. Please login again."
    log_application_error(msg, exception=error, status_code=401)
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({
            "success": False,
            "message": msg,
            "error_code": "UNAUTHORIZED",
            "details": str(error)
        }), 401
    return redirect(url_for('login'))

@app.errorhandler(403)
def forbidden_error(error):
    msg = "You do not have permission to access this resource."
    log_application_error(msg, exception=error, status_code=403)
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({
            "success": False,
            "message": msg,
            "error_code": "FORBIDDEN",
            "details": str(error)
        }), 403
    return render_template('403.html', message=msg), 403

@app.errorhandler(404)
def not_found_error(error):
    msg = "The requested resource could not be found."
    log_application_error(msg, exception=error, status_code=404)
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({
            "success": False,
            "message": msg,
            "error_code": "NOT_FOUND",
            "details": str(error)
        }), 404
    return render_template('403.html', message=msg), 404

@app.errorhandler(405)
def method_not_allowed_error(error):
    msg = "HTTP Method not allowed for this route."
    log_application_error(msg, exception=error, status_code=405)
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({
            "success": False,
            "message": msg,
            "error_code": "METHOD_NOT_ALLOWED",
            "details": str(error)
        }), 405
    return render_template('403.html', message=msg), 405

@app.errorhandler(500)
@app.errorhandler(Exception)
def internal_server_error(error):
    msg = "An internal server error occurred. Please try again later."
    log_application_error(f"Unhandled Exception: {str(error)}", exception=error, status_code=500)
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({
            "success": False,
            "message": msg,
            "error_code": "INTERNAL_SERVER_ERROR",
            "details": str(error)
        }), 500
    return render_template('500.html', message=msg), 500

@app.route('/api/customers', methods=['GET'])
@login_required
@permission_required('customer_history_view')
def api_get_customers():
    customers = load_customers()
    return jsonify({"success": True, "data": customers})

@app.route('/api/customers/<customer_id>/verify-billing', methods=['POST'])
@login_required
@permission_required('billing_verify')
def api_verify_billing(customer_id):
    data = request.get_json(silent=True) or {}
    admin_remarks = data.get('admin_remarks', '').strip()

    customers = load_customers()
    customer_idx = next((i for i, c in enumerate(customers) if c.get('customer_id') == customer_id), -1)
    if customer_idx == -1:
        return jsonify({"success": False, "message": "Customer record not found."}), 404

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    emp_id = session.get('employee_id', 'unknown')

    customers[customer_idx].update({
        "billing_status": "Billing Verified",
        "billing_verified": True,
        "billing_verified_by": emp_id,
        "billing_verified_at": now_str,
        "billing_admin_remarks": admin_remarks or None,
        "record_locked": True,
        "locked_by": emp_id,
        "locked_at": now_str
    })

    save_customers(customers)

    # Log to audit log
    db_add_audit_log({
        "action": f"verify_billing",
        "performed_by": session.get('username', 'Unknown'),
        "employee_id": emp_id,
        "role": session.get('role', 'admin'),
        "timestamp": now_str,
        "records_deleted": 0  # Conform to existing schema
    })

    return jsonify({
        "success": True, 
        "message": "Billing successfully verified and customer record locked.",
        "customer": customers[customer_idx]
    })

@app.route('/api/customers/<customer_id>/reopen', methods=['POST'])
@login_required
@permission_required('billing_reopen')
def api_reopen_billing(customer_id):
    data = request.get_json(silent=True) or {}
    reopen_reason = data.get('reopen_reason', '').strip()

    if not reopen_reason:
        return jsonify({"success": False, "message": "Reason for reopening billing is required."}), 400

    customers = load_customers()
    customer_idx = next((i for i, c in enumerate(customers) if c.get('customer_id') == customer_id), -1)
    if customer_idx == -1:
        return jsonify({"success": False, "message": "Customer record not found."}), 404

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    emp_id = session.get('employee_id', 'unknown')

    customers[customer_idx].update({
        "billing_status": "Pending Verification",
        "billing_verified": False,
        "record_locked": False,
        "reopened_by": emp_id,
        "reopened_at": now_str,
        "reopen_reason": reopen_reason
    })

    save_customers(customers)

    # Log to audit log
    db_add_audit_log({
        "action": f"reopen_billing",
        "performed_by": session.get('username', 'Unknown'),
        "employee_id": emp_id,
        "role": session.get('role', 'super_admin'),
        "timestamp": now_str,
        "records_deleted": 0  # Conform to existing schema
    })

    return jsonify({
        "success": True, 
        "message": "Billing reopened successfully and customer record unlocked.",
        "customer": customers[customer_idx]
    })

# --- SUPER ADMIN EXPORT ENDPOINTS ---

def get_filtered_customers(scope, search, from_date, to_date):
    customers = load_customers()
    
    # 1. Scope filter
    if scope == 'today':
        today_str = datetime.now().strftime('%Y-%m-%d')
        customers = [c for c in customers if c.get('created_at') and c.get('created_at').startswith(today_str)]
    elif scope == 'finance':
        customers = [c for c in customers if c.get('transaction_mode') == 'Finance']
        
    # 2. Date filters
    if from_date:
        customers = [c for c in customers if c.get('created_at') and c.get('created_at').split(' ')[0] >= from_date]
    if to_date:
        customers = [c for c in customers if c.get('created_at') and c.get('created_at').split(' ')[0] <= to_date]
        
    # 3. Search query filter
    if search:
        search = search.strip().lower()
        customers = [
            c for c in customers if
            search in (c.get('customer_name') or '').lower() or
            search in (c.get('mobile_number') or '') or
            search in (c.get('imei_number') or '') or
            search in (c.get('item_model') or '').lower() or
            search in (c.get('sales_person') or '').lower()
        ]
        
    return customers

def log_export_action(export_format, scope, filters, record_count):
    try:
        db_add_audit_log({
            "action": f"export_{export_format}_{scope}",
            "performed_by": session.get('username', 'Unknown'),
            "employee_id": session.get('employee_id', 'Unknown'),
            "role": session.get('role', 'super_admin'),
            "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "records_deleted": record_count
        })
    except Exception as e:
        print(f"Failed to write export audit log: {e}")

COLUMNS_TO_EXPORT = [
    "Customer ID", "Customer Name", "Mobile Number", "Item / Model", "IMEI Number",
    "Transaction Mode", "Finance Provider", "Down Payment Mode", "Down Payment Value",
    "Cash Amount", "Card Amount", "E-Wallet Amount", "Total Amount Received",
    "Exchange Status", "Exchange Brand", "Exchange Value", "Sales Person",
    "Remarks", "Created By", "Created At", "Updated At",
    "Billing Status", "Billing Verified", "Billing Verified By", "Billing Verified At",
    "Billing Admin Remarks", "Record Locked", "Locked By", "Locked At",
    "Reopened By", "Reopened At", "Reopen Reason"
]

KEY_MAPPING = {
    "Customer ID": "customer_id",
    "Customer Name": "customer_name",
    "Mobile Number": "mobile_number",
    "Item / Model": "item_model",
    "IMEI Number": "imei_number",
    "Transaction Mode": "transaction_mode",
    "Finance Provider": "finance_provider",
    "Down Payment Mode": "down_payment_mode",
    "Down Payment Value": "down_payment_value",
    "Cash Amount": "cash_amount",
    "Card Amount": "card_amount",
    "E-Wallet Amount": "ewallet_amount",
    "Total Amount Received": "total_amount_received",
    "Exchange Status": "exchange_status",
    "Exchange Brand": "exchange_brand",
    "Exchange Value": "exchange_value",
    "Sales Person": "sales_person",
    "Remarks": "remarks",
    "Created By": "created_by",
    "Created At": "created_at",
    "Updated At": "updated_at",
    "Billing Status": "billing_status",
    "Billing Verified": "billing_verified",
    "Billing Verified By": "billing_verified_by",
    "Billing Verified At": "billing_verified_at",
    "Billing Admin Remarks": "billing_admin_remarks",
    "Record Locked": "record_locked",
    "Locked By": "locked_by",
    "Locked At": "locked_at",
    "Reopened By": "reopened_by",
    "Reopened At": "reopened_at",
    "Reopen Reason": "reopen_reason"
}

@app.route('/api/export/customers/excel', methods=['GET'])
@login_required
@permission_required('export_data')
def api_export_customers_excel():
    try:
        scope = request.args.get('scope', 'all').strip().lower()
        search = request.args.get('search', '').strip()
        from_date = request.args.get('from_date', '').strip()
        to_date = request.args.get('to_date', '').strip()

        customers = get_filtered_customers(scope, search, from_date, to_date)
        
        # Generate Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"
        
        # Append Header
        ws.append(COLUMNS_TO_EXPORT)
        
        # Style Header
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid") # green theme
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        for col_num in range(1, len(COLUMNS_TO_EXPORT) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Append rows
        for customer in customers:
            row = []
            for col in COLUMNS_TO_EXPORT:
                key = KEY_MAPPING[col]
                row.append(customer.get(key, ''))
            ws.append(row)
            
        # Auto column width adjustment
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        today_date = datetime.now().strftime('%Y-%m-%d')
        filename = f"customers_{scope}_{today_date}.xlsx"
        
        log_export_action('excel', scope, {'search': search, 'from_date': from_date, 'to_date': to_date}, len(customers))
        
        return send_file(
            file_stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"success": False, "message": f"Failed to export Excel: {str(e)}"}), 500

@app.route('/api/export/customers/csv', methods=['GET'])
@login_required
@permission_required('export_data')
def api_export_customers_csv():
    try:
        scope = request.args.get('scope', 'all').strip().lower()
        search = request.args.get('search', '').strip()
        from_date = request.args.get('from_date', '').strip()
        to_date = request.args.get('to_date', '').strip()

        customers = get_filtered_customers(scope, search, from_date, to_date)
        
        # Generate CSV
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(COLUMNS_TO_EXPORT)
        
        for customer in customers:
            row = []
            for col in COLUMNS_TO_EXPORT:
                key = KEY_MAPPING[col]
                row.append(customer.get(key, ''))
            writer.writerow(row)
            
        mem = io.BytesIO()
        mem.write(si.getvalue().encode('utf-8'))
        mem.seek(0)
        
        today_date = datetime.now().strftime('%Y-%m-%d')
        filename = f"customers_{scope}_{today_date}.csv"
        
        log_export_action('csv', scope, {'search': search, 'from_date': from_date, 'to_date': to_date}, len(customers))
        
        return send_file(
            mem,
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"success": False, "message": f"Failed to export CSV: {str(e)}"}), 500

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 7)
        self.setFillColor(colors.HexColor("#475569"))
        
        # Header (on all pages)
        self.drawString(15, 560, "Sangeetha Customer Records Report")
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(15, 552, 827, 552)
        
        # Footer (on all pages)
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(827, 20, page_text)
        self.drawString(15, 20, "Confidential - For Internal Use Only")
        self.restoreState()

@app.route('/api/export/customers/pdf', methods=['GET'])
@login_required
@permission_required('export_data')
def api_export_customers_pdf():
    try:
        scope = request.args.get('scope', 'all').strip().lower()
        search = request.args.get('search', '').strip()
        from_date = request.args.get('from_date', '').strip()
        to_date = request.args.get('to_date', '').strip()

        customers = get_filtered_customers(scope, search, from_date, to_date)
        
        # Generate PDF
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(A4),
            leftMargin=15,
            rightMargin=15,
            topMargin=55,
            bottomMargin=45
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # Title Styling
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=16,
            leading=20,
            textColor=colors.HexColor("#0D70C0"),
            spaceAfter=12
        )
        
        metadata_style = ParagraphStyle(
            'ReportMetadata',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#475569")
        )
        
        # Banner Title
        story.append(Paragraph("Sangeetha Customer Records Report", title_style))
        
        # Metadata Block
        gen_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        user_name = session.get('username', 'Unknown')
        filter_str = f"Scope: {scope.capitalize()}"
        if search:
            filter_str += f" | Search: \"{search}\""
        if from_date or to_date:
            filter_str += f" | Range: {from_date or 'Start'} to {to_date or 'End'}"
            
        story.append(Paragraph(f"<strong>Generated At:</strong> {gen_time} &nbsp;&nbsp;|&nbsp;&nbsp; <strong>Exported By:</strong> {user_name} &nbsp;&nbsp;|&nbsp;&nbsp; <strong>Filters:</strong> {filter_str}", metadata_style))
        story.append(Spacer(1, 15))
        
        # Table Styling
        cell_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=5.5,
            leading=6.5,
            textColor=colors.HexColor("#1e293b")
        )
        header_cell_style = ParagraphStyle(
            'TableHeaderCell',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=6,
            leading=7.5,
            textColor=colors.white,
            alignment=1 # Center
        )
        
        table_data = []
        # Header Row
        header_row = [Paragraph(col, header_cell_style) for col in COLUMNS_TO_EXPORT]
        table_data.append(header_row)
        
        # Data Rows
        for customer in customers:
            data_row = []
            for col in COLUMNS_TO_EXPORT:
                key = KEY_MAPPING[col]
                val = str(customer.get(key, '') or '')
                data_row.append(Paragraph(val, cell_style))
            table_data.append(data_row)
            
        # exact column widths to sum up to exactly 812 pt
        col_widths = [25, 34, 35, 35, 35, 30, 30, 25, 25, 20, 20, 20, 25, 18, 22, 20, 30, 25, 25, 26, 26, 25, 18, 25, 25, 25, 18, 25, 25, 25, 25, 25]
        
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#DC2626")), # red theme header
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]), # Alternating
        ]))
        
        story.append(t)
        
        # Build Document
        doc.build(story, canvasmaker=NumberedCanvas)
        
        pdf_buffer.seek(0)
        today_date = datetime.now().strftime('%Y-%m-%d')
        filename = f"customers_{scope}_{today_date}.pdf"
        
        log_export_action('pdf', scope, {'search': search, 'from_date': from_date, 'to_date': to_date}, len(customers))
        
        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"success": False, "message": f"Failed to export PDF: {str(e)}"}), 500

@app.route('/api/verify-imei', methods=['POST'])
@login_required
@permission_required('customer_create')
def api_verify_imei():
    data = request.get_json() or {}
    imei = data.get('imei', '').strip()

    if not imei:
        return jsonify({"success": False, "message": "IMEI is required."}), 400

    return jsonify({"success": True, "message": "IMEI verified successfully."})

@app.route('/api/customers', methods=['POST'])
@login_required
@permission_required('customer_create')
def api_add_customer():
    data = request.get_json() or {}

    # Extract fields
    customer_name = data.get('customer_name', '').strip()
    mobile_number = data.get('mobile_number', '').strip()
    transaction_mode = data.get('transaction_mode', '').strip()
    finance_provider = data.get('finance_provider', '').strip()
    down_payment_mode = data.get('down_payment_mode', '').strip()
    down_payment_value_raw = data.get('down_payment_value', 0)
    
    cash_amount_raw = data.get('cash_amount', 0)
    card_amount_raw = data.get('card_amount', 0)
    ewallet_amount_raw = data.get('ewallet_amount', 0)
    
    item_model = data.get('item_model', '').strip()
    imei_number = data.get('imei_number', '').strip()
    
    exchange_status = data.get('exchange_status', '').strip()
    exchange_brand = data.get('exchange_brand', '').strip()
    exchange_value_raw = data.get('exchange_value', 0)
    
    sales_person = data.get('sales_person', '').strip()
    remarks = data.get('remarks', '').strip()

    # Name validations
    if not customer_name:
        return jsonify({"success": False, "message": "Customer Name is required."}), 400
    if len(customer_name) > 100:
        return jsonify({"success": False, "message": "Customer Name must not exceed 100 characters."}), 400
    customer_name = ' '.join(word.capitalize() for word in customer_name.split())

    # Mobile validations
    if not mobile_number:
        return jsonify({"success": False, "message": "Mobile Number is required."}), 400
    if len(mobile_number) != 10 or not mobile_number.isdigit():
        return jsonify({"success": False, "message": "Mobile Number must be exactly 10 digits."}), 400

    # Model & IMEI validations
    if not item_model:
        return jsonify({"success": False, "message": "Item / Model is required."}), 400
    if len(item_model) > 200:
        return jsonify({"success": False, "message": "Item / Model must not exceed 200 characters."}), 400

    if not imei_number:
        return jsonify({"success": False, "message": "IMEI Number is required."}), 400

    # Transaction Mode validation
    if transaction_mode not in ['Finance', 'Non-Finance', 'Non Finance']:
        return jsonify({"success": False, "message": "Transaction Mode must be Finance or Non-Finance."}), 400

    # Finance specific validation
    if transaction_mode == 'Finance':
        allowed_providers = [
            'Bajaj Finserv Lending Financier', 'Bajaj Finservlending Financier', 'Bajaj',
            'DMI Consumer Credit (Samsung, Oppo, Vivo)', 'DMI Consumer Credit (Samsung , Oppo, Vivo)', 'DMI-OPPO', 'DMI-VIVO', 'DMI-other',
            'HDB Financial Services Ltd (Financier)', 'HDB',
            'IDFC First Bank Ltd (Financier)', 'IDFC',
            'TVS Credit Services Limited (Financier)', 'TVS Credit Services Limited(Financier)', 'TVS'
        ]
        if not finance_provider:
            return jsonify({"success": False, "message": "Finance Provider is required when Transaction Mode is Finance."}), 400
        if finance_provider not in allowed_providers:
            return jsonify({"success": False, "message": "Invalid Finance Provider selected."}), 400

        if down_payment_mode not in ['Cash', 'Card', 'E-Wallet']:
            return jsonify({"success": False, "message": "Down Payment Mode is required and must be Cash, Card, or E-Wallet."}), 400

        try:
            down_payment_value = float(down_payment_value_raw)
            if down_payment_value < 0:
                raise ValueError
        except ValueError:
            return jsonify({"success": False, "message": "Down Payment Value must be a valid positive number."}), 400

        # Zero out payment split fields in Finance mode
        cash_amount = 0.0
        card_amount = 0.0
        ewallet_amount = 0.0
        total_amount_received = down_payment_value

    else:
        # Non-Finance specific validations
        finance_provider = ""
        down_payment_mode = ""
        down_payment_value = 0.0

        try:
            cash_amount = float(cash_amount_raw) if cash_amount_raw else 0.0
            card_amount = float(card_amount_raw) if card_amount_raw else 0.0
            ewallet_amount = float(ewallet_amount_raw) if ewallet_amount_raw else 0.0
            if cash_amount < 0 or card_amount < 0 or ewallet_amount < 0:
                raise ValueError
        except ValueError:
            return jsonify({"success": False, "message": "Payment split amounts must be positive numbers."}), 400

        total_amount_received = round(cash_amount + card_amount + ewallet_amount, 2)
        if total_amount_received <= 0:
            return jsonify({"success": False, "message": "At least one payment amount (Cash, Card, or E-Wallet) must be greater than zero."}), 400

    # Exchange validations
    if exchange_status not in ['Yes', 'No']:
        return jsonify({"success": False, "message": "Exchange Status must be Yes or No."}), 400

    if exchange_status == 'Yes':
        if not exchange_brand:
            return jsonify({"success": False, "message": "Exchange Brand is required when Exchange is Yes."}), 400
        try:
            exchange_value = float(exchange_value_raw)
            if exchange_value < 0:
                raise ValueError
        except ValueError:
            return jsonify({"success": False, "message": "Exchange Value must be a valid positive number."}), 400
    else:
        exchange_brand = ""
        exchange_value = 0.0

    # Salesperson validation
    allowed_salespeople = get_allowed_salespeople()
    if not sales_person or sales_person not in allowed_salespeople:
        return jsonify({"success": False, "message": "Invalid Sales Person selected."}), 400

    # Remarks validation
    if len(remarks) > 500:
        return jsonify({"success": False, "message": "Remarks must not exceed 500 characters."}), 400

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Generate customer_id CUST-XXXX
    customers = load_customers()
    next_num = 1001
    if customers:
        for c in customers:
            cid = c.get('customer_id', '')
            if cid.startswith('CUST-'):
                try:
                    num = int(cid.split('-')[1])
                    if num >= next_num:
                        next_num = num + 1
                except (ValueError, IndexError):
                    pass
    customer_id = f"CUST-{next_num}"

    new_customer = {
        "customer_id": customer_id,
        "customer_name": customer_name,
        "mobile_number": mobile_number,
        "item_model": item_model,
        "imei_number": imei_number,
        "transaction_mode": 'Finance' if transaction_mode == 'Finance' else 'Non-Finance',
        "finance_provider": finance_provider,
        "down_payment_mode": down_payment_mode,
        "down_payment_value": down_payment_value,
        "cash_amount": cash_amount,
        "card_amount": card_amount,
        "ewallet_amount": ewallet_amount,
        "total_amount_received": total_amount_received,
        "exchange_status": exchange_status,
        "exchange_brand": exchange_brand,
        "exchange_value": exchange_value,
        "sales_person": sales_person,
        "remarks": remarks,
        "created_by": session.get('employee_id', 'admin'),
        "created_at": now_str,
        "updated_at": now_str,
        "billing_status": "Pending Verification",
        "billing_verified": False,
        "billing_verified_by": None,
        "billing_verified_at": None,
        "billing_admin_remarks": None,
        "record_locked": False,
        "locked_by": None,
        "locked_at": None,
        "reopened_by": None,
        "reopened_at": None,
        "reopen_reason": None
    }

    customers.append(new_customer)
    try:
        save_customers(customers)
    except Exception as e:
        print(f"ERROR saving customer to database: {e}")
        return jsonify({"success": False, "message": f"Database save failed: {str(e)}"}), 500

    return jsonify({"success": True, "message": "Customer record saved successfully!", "customer": new_customer})


@app.route('/api/raw-data', methods=['GET'])
@login_required
@permission_required('settings_access')
def api_get_raw_data():
    file_param = request.args.get('file', 'customers.json')
    if file_param not in WHITELIST_FILES:
        return jsonify({"success": False, "message": "Invalid file parameter"}), 400
        
    try:
        if file_param == 'customers.json':
            data = load_customers()
        elif file_param == 'crm_walkin_customers.json':
            data = load_crm_customers()
        elif file_param == 'customer_edit_requests.json':
            data = load_edit_requests()
        elif file_param == 'users.json':
            data = load_users()
        elif file_param == 'payments.json':
            data = load_payments()
        else:
            data = []
        return jsonify({"success": True, "raw": json.dumps(data, indent=4)})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/raw-data', methods=['POST'])
@login_required
@permission_required('settings_access')
def api_save_raw_data():
    file_param = request.args.get('file', 'customers.json')
    if file_param not in WHITELIST_FILES:
        return jsonify({"success": False, "message": "Invalid file parameter"}), 400
        
    data = request.get_json() or {}
    raw_text = data.get('raw', '').strip()
    
    if not raw_text:
        return jsonify({"success": False, "message": "Raw JSON content is empty."}), 400
        
    try:
        parsed_json = json.loads(raw_text)
        if not isinstance(parsed_json, list):
            return jsonify({"success": False, "message": "JSON must be a list (array) of objects."}), 400
            
        if file_param == 'customers.json':
            save_customers(parsed_json)
        elif file_param == 'crm_walkin_customers.json':
            save_crm_customers(parsed_json)
        elif file_param == 'customer_edit_requests.json':
            save_edit_requests(parsed_json)
        elif file_param == 'users.json':
            save_users(parsed_json)
        elif file_param == 'payments.json':
            save_payments(parsed_json)
            
        return jsonify({"success": True, "message": f"{file_param} database file saved successfully!"})
    except json.JSONDecodeError as err:
        return jsonify({"success": False, "message": f"Invalid JSON syntax: {str(err)}"}), 400
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/customers/<customer_id>', methods=['DELETE'])
@login_required
@permission_required('customer_history_delete')
def api_delete_customer(customer_id):
    role = session.get('role')
    if role not in ['super_admin', 'admin']:
        return jsonify({"success": False, "message": "Unauthorized"}), 403

    customers = load_customers()
    customer = next((c for c in customers if c.get('customer_id') == customer_id), None)
    if not customer:
        return jsonify({"success": False, "message": "Record not found."}), 404
        
    if customer.get('record_locked') and role != 'super_admin':
        return jsonify({"success": False, "message": "This customer record is locked (billing verified) and cannot be deleted."}), 403
        
    try:
        db_delete_customer_record(customer_id)
        return jsonify({"success": True, "message": "Customer record permanently deleted from database and storage!"})
    except Exception as e:
        log_application_error("Failed to delete customer record", e, 500)
        return jsonify({"success": False, "message": f"Deletion failed: {str(e)}"}), 500

@app.route('/api/customers/clear-all', methods=['POST'])
@login_required
@permission_required('customer_clear_all')
def api_clear_all_customers():
    """Clear ALL customer records. Preserves all other data (users, payments, edit_requests)."""
    customers = load_customers()
    records_deleted = len(customers)

    # Write audit log entry to Supabase
    try:
        db_add_audit_log({
            "action": "clear_all_customers",
            "performed_by": session.get('username', ''),
            "employee_id": session.get('employee_id', ''),
            "role": session.get('role', ''),
            "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "records_deleted": records_deleted
        })
    except Exception:
        pass  # Audit log failure should not block the operation


    # Clear all customers
    save_customers([])
    return jsonify({
        "success": True,
        "message": f"All {records_deleted} customer record(s) have been cleared successfully.",
        "records_deleted": records_deleted
    })


# --- CRM WALKIN CUSTOMERS SECTION ---

@app.route('/crm-walkin-customers')
@login_required
@permission_required('crm_view')
def crm_walkin_customers_view():
    return render_template('crm_walkin_customers.html', username=session['username'], role=session['role'])

@app.route('/api/crm-walkin-customers', methods=['GET'])
@login_required
@permission_required('crm_view')
def api_get_crm_customers():
    customers = load_crm_customers()
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    reason = request.args.get('reason')
    sales_person = request.args.get('sales_person')
    search = request.args.get('search')
    
    filtered = []
    for c in customers:
        created_date = c.get('created_at', '').split(' ')[0]
        if start_date and created_date < start_date:
            continue
        if end_date and created_date > end_date:
            continue
        if reason and c.get('walkout_reason') != reason:
            continue
        if sales_person and c.get('sales_person') != sales_person:
            continue
        if search:
            s_lower = search.lower()
            name = c.get('customer_name', '').lower()
            mobile = c.get('mobile_number', '').lower()
            model = c.get('model_item', '').lower()
            if s_lower not in name and s_lower not in mobile and s_lower not in model:
                continue
        filtered.append(c)
        
    filtered.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return jsonify({"success": True, "data": filtered})

@app.route('/api/crm-walkin-customers', methods=['POST'])
@login_required
@permission_required('crm_create')
def api_add_crm_customer():
    data = request.get_json() or {}
    
    customer_name = data.get('customer_name', '').strip()
    mobile_number = data.get('mobile_number', '').strip()
    model_item = data.get('model_item', '').strip()
    walkout_reason = data.get('walkout_reason', '').strip()
    remarks = data.get('remarks', '').strip()
    sales_person = data.get('sales_person', '').strip()
    
    if not customer_name:
        return jsonify({"success": False, "message": "Customer Name is required."}), 400
    if len(customer_name) > 100:
        return jsonify({"success": False, "message": "Customer Name must not exceed 100 characters."}), 400
        
    if not mobile_number:
        return jsonify({"success": False, "message": "Mobile Number is required."}), 400
    if len(mobile_number) != 10 or not mobile_number.isdigit():
        return jsonify({"success": False, "message": "Mobile Number must be exactly 10 numeric digits."}), 400
        
    if not model_item:
        return jsonify({"success": False, "message": "Model / Item Enquired is required."}), 400
    if len(model_item) > 200:
        return jsonify({"success": False, "message": "Model / Item Enquired must not exceed 200 characters."}), 400
        
    allowed_reasons = [
        "No Walkout - Purchasing at Sangeetha",
        "Just Walk-in Customer",
        "Enquiry Customer",
        "Pricing Issue",
        "others"
    ]
    if not walkout_reason or walkout_reason not in allowed_reasons:
        return jsonify({"success": False, "message": "Invalid or missing Walk-in / Walkout Reason."}), 400
        
    allowed_salespeople = get_allowed_salespeople()
    if not sales_person or sales_person not in allowed_salespeople:
        return jsonify({"success": False, "message": "Invalid or missing Sales Person."}), 400
        
    if len(remarks) > 500:
        return jsonify({"success": False, "message": "Remarks must not exceed 500 characters."}), 400
        
    crm_customers = load_crm_customers()
    next_num = 1001
    if crm_customers:
        for c in crm_customers:
            cid = c.get('crm_customer_id', '')
            if cid.startswith('CRM-'):
                try:
                    num = int(cid.split('-')[1])
                    if num >= next_num:
                        next_num = num + 1
                except (ValueError, IndexError):
                    pass
    crm_customer_id = f"CRM-{next_num}"
    
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    new_record = {
        "crm_customer_id": crm_customer_id,
        "customer_name": ' '.join(w.capitalize() for w in customer_name.split()),
        "mobile_number": mobile_number,
        "model_item": model_item,
        "walkout_reason": walkout_reason,
        "remarks": remarks,
        "sales_person": sales_person,
        "created_by": session.get('employee_id', 'admin'),
        "created_at": now_str,
        "updated_at": now_str
    }
    
    crm_customers.append(new_record)
    save_crm_customers(crm_customers)
    return jsonify({"success": True, "message": "CRM record saved successfully!", "data": new_record})

@app.route('/api/crm-walkin-customers/<crm_customer_id>', methods=['PUT'])
@login_required
@permission_required('crm_edit')
def api_update_crm_customer(crm_customer_id):
    data = request.get_json() or {}
    
    customer_name = data.get('customer_name', '').strip()
    mobile_number = data.get('mobile_number', '').strip()
    model_item = data.get('model_item', '').strip()
    walkout_reason = data.get('walkout_reason', '').strip()
    remarks = data.get('remarks', '').strip()
    sales_person = data.get('sales_person', '').strip()
    
    if not customer_name:
        return jsonify({"success": False, "message": "Customer Name is required."}), 400
    if len(customer_name) > 100:
        return jsonify({"success": False, "message": "Customer Name must not exceed 100 characters."}), 400
        
    if not mobile_number:
        return jsonify({"success": False, "message": "Mobile Number is required."}), 400
    if len(mobile_number) != 10 or not mobile_number.isdigit():
        return jsonify({"success": False, "message": "Mobile Number must be exactly 10 numeric digits."}), 400
        
    if not model_item:
        return jsonify({"success": False, "message": "Model / Item Enquired is required."}), 400
    if len(model_item) > 200:
        return jsonify({"success": False, "message": "Model / Item Enquired must not exceed 200 characters."}), 400
        
    allowed_reasons = [
        "No Walkout - Purchasing at Sangeetha",
        "Just Walk-in Customer",
        "Enquiry Customer",
        "Pricing Issue",
        "others"
    ]
    if not walkout_reason or walkout_reason not in allowed_reasons:
        return jsonify({"success": False, "message": "Invalid or missing Walk-in / Walkout Reason."}), 400
        
    allowed_salespeople = get_allowed_salespeople()
    if not sales_person or sales_person not in allowed_salespeople:
        return jsonify({"success": False, "message": "Invalid or missing Sales Person."}), 400
        
    if len(remarks) > 500:
        return jsonify({"success": False, "message": "Remarks must not exceed 500 characters."}), 400
        
    crm_customers = load_crm_customers()
    idx = next((i for i, c in enumerate(crm_customers) if c.get('crm_customer_id') == crm_customer_id), -1)
    if idx == -1:
        return jsonify({"success": False, "message": "CRM record not found."}), 404
        
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    crm_customers[idx].update({
        "customer_name": ' '.join(w.capitalize() for w in customer_name.split()),
        "mobile_number": mobile_number,
        "model_item": model_item,
        "walkout_reason": walkout_reason,
        "remarks": remarks,
        "sales_person": sales_person,
        "updated_at": now_str
    })
    
    save_crm_customers(crm_customers)
    return jsonify({"success": True, "message": "CRM record updated successfully!", "data": crm_customers[idx]})

@app.route('/api/crm-walkin-customers/<crm_customer_id>', methods=['DELETE'])
@login_required
@permission_required('crm_delete')
def api_delete_crm_customer(crm_customer_id):
    crm_customers = load_crm_customers()
    updated = [c for c in crm_customers if c.get('crm_customer_id') != crm_customer_id]
    
    if len(updated) == len(crm_customers):
        return jsonify({"success": False, "message": "CRM record not found."}), 404
        
    save_crm_customers(updated)
    return jsonify({"success": True, "message": "CRM record deleted successfully!"})


@app.route('/customers/<customer_id>/request-edit')
@login_required
def edit_request_view(customer_id):
    # Check if the user role has permission to submit edit request or history edit
    role = session.get('role')
    allowed_permissions = ROLE_PERMISSIONS.get(role, [])
    if 'customer_edit_request_create' not in allowed_permissions and 'customer_history_edit' not in allowed_permissions:
        return render_template('403.html', message="You do not have permission to edit customer records."), 403

    customers = load_customers()
    customer = next((c for c in customers if c.get('customer_id') == customer_id), None)
    if not customer:
        return render_template('403.html', message="Customer record not found."), 404
    return render_template('edit_request.html', customer=customer, username=session['username'], role=session['role'])

@app.route('/api/customers/<customer_id>', methods=['PUT'])
@login_required
@permission_required('customer_history_edit')
def api_update_customer(customer_id):
    customers = load_customers()
    customer_index = next((i for i, c in enumerate(customers) if c.get('customer_id') == customer_id), -1)
    if customer_index == -1:
        return jsonify({"success": False, "message": "Record not found."}), 404
    if customers[customer_index].get('record_locked'):
        return jsonify({"success": False, "message": "This customer record is locked (billing verified) and cannot be updated."}), 403

    data = request.get_json() or {}
    
    # Extract fields
    customer_name = data.get('customer_name', '').strip()
    mobile_number = data.get('mobile_number', '').strip()
    transaction_mode = data.get('transaction_mode', '').strip()
    finance_provider = data.get('finance_provider', '').strip()
    down_payment_mode = data.get('down_payment_mode', '').strip()
    down_payment_value_raw = data.get('down_payment_value', 0)
    
    cash_amount_raw = data.get('cash_amount', 0)
    card_amount_raw = data.get('card_amount', 0)
    ewallet_amount_raw = data.get('ewallet_amount', 0)
    
    item_model = data.get('item_model', '').strip()
    imei_number = data.get('imei_number', '').strip()
    
    exchange_status = data.get('exchange_status', '').strip()
    exchange_brand = data.get('exchange_brand', '').strip()
    exchange_value_raw = data.get('exchange_value', 0)
    
    sales_person = data.get('sales_person', '').strip()
    remarks = data.get('remarks', '').strip()

    # Name validations
    if not customer_name:
        return jsonify({"success": False, "message": "Customer Name is required."}), 400
    if len(customer_name) > 100:
        return jsonify({"success": False, "message": "Customer Name must not exceed 100 characters."}), 400
    customer_name = ' '.join(word.capitalize() for word in customer_name.split())

    # Mobile validations
    if not mobile_number:
        return jsonify({"success": False, "message": "Mobile Number is required."}), 400
    if len(mobile_number) != 10 or not mobile_number.isdigit():
        return jsonify({"success": False, "message": "Mobile Number must be exactly 10 digits."}), 400

    # Model & IMEI validations
    if not item_model:
        return jsonify({"success": False, "message": "Item / Model is required."}), 400
    if len(item_model) > 200:
        return jsonify({"success": False, "message": "Item / Model must not exceed 200 characters."}), 400

    if not imei_number:
        return jsonify({"success": False, "message": "IMEI Number is required."}), 400

    # Transaction Mode validation
    if transaction_mode not in ['Finance', 'Non-Finance', 'Non Finance']:
        return jsonify({"success": False, "message": "Transaction Mode must be Finance or Non-Finance."}), 400

    # Finance specific validation
    if transaction_mode == 'Finance':
        allowed_providers = [
            'Bajaj Finserv Lending Financier', 'Bajaj Finservlending Financier', 'Bajaj',
            'DMI Consumer Credit (Samsung, Oppo, Vivo)', 'DMI Consumer Credit (Samsung , Oppo, Vivo)', 'DMI-OPPO', 'DMI-VIVO', 'DMI-other',
            'HDB Financial Services Ltd (Financier)', 'HDB',
            'IDFC First Bank Ltd (Financier)', 'IDFC',
            'TVS Credit Services Limited (Financier)', 'TVS Credit Services Limited(Financier)', 'TVS'
        ]
        if not finance_provider:
            return jsonify({"success": False, "message": "Finance Provider is required when Transaction Mode is Finance."}), 400
        if finance_provider not in allowed_providers:
            return jsonify({"success": False, "message": "Invalid Finance Provider selected."}), 400

        if down_payment_mode not in ['Cash', 'Card', 'E-Wallet']:
            return jsonify({"success": False, "message": "Down Payment Mode is required and must be Cash, Card, or E-Wallet."}), 400

        try:
            down_payment_value = float(down_payment_value_raw)
            if down_payment_value < 0:
                raise ValueError
        except ValueError:
            return jsonify({"success": False, "message": "Down Payment Value must be a valid positive number."}), 400

        # Zero out payment split fields in Finance mode
        cash_amount = 0.0
        card_amount = 0.0
        ewallet_amount = 0.0
        total_amount_received = down_payment_value

    else:
        # Non-Finance specific validations
        finance_provider = ""
        down_payment_mode = ""
        down_payment_value = 0.0

        try:
            cash_amount = float(cash_amount_raw) if cash_amount_raw else 0.0
            card_amount = float(card_amount_raw) if card_amount_raw else 0.0
            ewallet_amount = float(ewallet_amount_raw) if ewallet_amount_raw else 0.0
            if cash_amount < 0 or card_amount < 0 or ewallet_amount < 0:
                raise ValueError
        except ValueError:
            return jsonify({"success": False, "message": "Payment split amounts must be positive numbers."}), 400

        total_amount_received = round(cash_amount + card_amount + ewallet_amount, 2)
        if total_amount_received <= 0:
            return jsonify({"success": False, "message": "At least one payment amount (Cash, Card, or E-Wallet) must be greater than zero."}), 400

    # Exchange validations
    if exchange_status not in ['Yes', 'No']:
        return jsonify({"success": False, "message": "Exchange Status must be Yes or No."}), 400

    if exchange_status == 'Yes':
        if not exchange_brand:
            return jsonify({"success": False, "message": "Exchange Brand is required when Exchange is Yes."}), 400
        try:
            exchange_value = float(exchange_value_raw)
            if exchange_value < 0:
                raise ValueError
        except ValueError:
            return jsonify({"success": False, "message": "Exchange Value must be a valid positive number."}), 400
    else:
        exchange_brand = ""
        exchange_value = 0.0

    # Salesperson validation
    allowed_salespeople = get_allowed_salespeople()
    if not sales_person or sales_person not in allowed_salespeople:
        return jsonify({"success": False, "message": "Invalid Sales Person selected."}), 400

    # Remarks validation
    if len(remarks) > 500:
        return jsonify({"success": False, "message": "Remarks must not exceed 500 characters."}), 400

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Update customer record
    customers[customer_index].update({
        "customer_name": customer_name,
        "mobile_number": mobile_number,
        "item_model": item_model,
        "imei_number": imei_number,
        "transaction_mode": 'Finance' if transaction_mode == 'Finance' else 'Non-Finance',
        "finance_provider": finance_provider,
        "down_payment_mode": down_payment_mode,
        "down_payment_value": down_payment_value,
        "cash_amount": cash_amount,
        "card_amount": card_amount,
        "ewallet_amount": ewallet_amount,
        "total_amount_received": total_amount_received,
        "exchange_status": exchange_status,
        "exchange_brand": exchange_brand,
        "exchange_value": exchange_value,
        "sales_person": sales_person,
        "remarks": remarks,
        "updated_at": now_str
    })

    save_customers(customers)
    return jsonify({"success": True, "message": "Customer record updated successfully!", "customer": customers[customer_index]})

def load_payments():
    """Load payment list from Supabase."""
    return db_load_payments()

def save_payments(payments_list):
    """Save payment list to Supabase."""
    db_save_payments(payments_list)


@app.route('/payment-tracker')
@login_required
@permission_required('payment_tracker_view')
def payment_tracker():
    return render_template('payment_tracker.html', username=session['username'], role=session['role'])

@app.route('/api/payment-summary', methods=['GET'])
@login_required
@permission_required('payment_tracker_view')
def api_payment_summary():
    payments = load_payments()
    today_str = datetime.now().strftime('%Y-%m-%d')
    
    total_sales = 0.0
    cash_sales = 0.0
    card_sales = 0.0
    ewallet_sales = 0.0
    total_received = 0.0
    total_pending = 0.0
    customer_pending = 0.0
    employee_pending = 0.0
    
    today_cash_sales = 0.0
    today_card_sales = 0.0
    today_ewallet_sales = 0.0
    today_pending = 0.0
    total_unsettled = 0.0
    
    for p in payments:
        total_bill = float(p.get('total_bill_amount') or 0)
        received = float(p.get('amount_received') or 0)
        pending = float(p.get('pending_amount') or 0)
        mode = p.get('payment_mode')
        pending_from = p.get('pending_from')
        created_at = p.get('created_at', '')
        
        total_sales += total_bill
        total_received += received
        total_pending += pending
        
        if mode == 'Cash':
            cash_sales += total_bill
        elif mode == 'Card':
            card_sales += total_bill
        elif mode == 'E-Wallet':
            ewallet_sales += total_bill
            
        if pending_from == 'Customer':
            customer_pending += pending
        elif pending_from == 'Employee / Salesperson':
            employee_pending += pending
            
        if p.get('settlement_status') == 'Unsettled':
            total_unsettled += pending
            
        # Today metrics
        if created_at.startswith(today_str):
            today_pending += pending
            if mode == 'Cash':
                today_cash_sales += total_bill
            elif mode == 'Card':
                today_card_sales += total_bill
            elif mode == 'E-Wallet':
                today_ewallet_sales += total_bill
                
    return jsonify({
        "success": True,
        "summary": {
            "total_sales_amount": total_sales,
            "cash_sales": cash_sales,
            "card_sales": card_sales,
            "ewallet_sales": ewallet_sales,
            "total_received": total_received,
            "total_pending": total_pending,
            "customer_pending": customer_pending,
            "employee_pending": employee_pending,
            "today_cash_sales": today_cash_sales,
            "today_card_sales": today_card_sales,
            "today_ewallet_sales": today_ewallet_sales,
            "today_pending_amount": today_pending,
            "total_unsettled_amount": total_unsettled
        }
    })

@app.route('/api/payments', methods=['GET'])
@login_required
@permission_required('payment_tracker_view')
def api_get_payments():
    payments = load_payments()
    
    # Apply query filters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    payment_mode = request.args.get('payment_mode')
    payment_status = request.args.get('payment_status')
    settlement_status = request.args.get('settlement_status')
    pending_from = request.args.get('pending_from')
    sales_person = request.args.get('sales_person')
    search = request.args.get('search')
    
    filtered_payments = []
    for p in payments:
        # Date range filter
        created_at_date = p.get('created_at', '').split(' ')[0]
        if start_date and created_at_date < start_date:
            continue
        if end_date and created_at_date > end_date:
            continue
            
        # Equality filters
        if payment_mode and p.get('payment_mode') != payment_mode:
            continue
        if payment_status and p.get('payment_status') != payment_status:
            continue
        if settlement_status and p.get('settlement_status') != settlement_status:
            continue
        if pending_from and p.get('pending_from') != pending_from:
            continue
        if sales_person and p.get('sales_person') != sales_person:
            continue
            
        # Search filter
        if search:
            search_lower = search.lower()
            c_name = p.get('customer_name', '').lower()
            mobile = p.get('mobile_number', '').lower()
            invoice = p.get('invoice_number', '').lower()
            imei = p.get('imei_number', '').lower()
            s_person = p.get('sales_person', '').lower()
            p_person = p.get('pending_person_name', '').lower()
            if (search_lower not in c_name and 
                search_lower not in mobile and 
                search_lower not in invoice and 
                search_lower not in imei and 
                search_lower not in s_person and 
                search_lower not in p_person):
                continue
                
        filtered_payments.append(p)
        
    # Sort payments: newer first
    filtered_payments.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return jsonify({"success": True, "data": filtered_payments})

@app.route('/api/payments', methods=['POST'])
@login_required
@permission_required('payment_tracker_create')
def api_add_payment():
    data = request.get_json() or {}
    
    # Extract fields
    customer_name = data.get('customer_name', '').strip()
    mobile_number = data.get('mobile_number', '').strip()
    invoice_number = data.get('invoice_number', '').strip()
    item_model = data.get('item_model', '').strip()
    imei_number = data.get('imei_number', '').strip()
    sales_person = data.get('sales_person', '').strip()
    payment_mode = data.get('payment_mode', '').strip()
    total_bill_amount = data.get('total_bill_amount')
    amount_received = data.get('amount_received')
    pending_from = data.get('pending_from', '').strip()
    pending_person_name = data.get('pending_person_name', '').strip()
    due_date = data.get('due_date', '').strip()
    remarks = data.get('remarks', '').strip()
    payment_status_override = data.get('payment_status', '').strip()
    settlement_status_override = data.get('settlement_status', '').strip()
    
    # Validation logic
    if not customer_name:
        return jsonify({"success": False, "message": "Customer Name is required."}), 400
    if len(customer_name) > 100:
        return jsonify({"success": False, "message": "Customer Name must not exceed 100 characters."}), 400
    customer_name = ' '.join(w.capitalize() for w in customer_name.split())

    if not mobile_number:
        return jsonify({"success": False, "message": "Mobile Number is required."}), 400
    if len(mobile_number) != 10 or not mobile_number.isdigit():
        return jsonify({"success": False, "message": "Mobile Number must be exactly 10 digits."}), 400

    if item_model and len(item_model) > 200:
        return jsonify({"success": False, "message": "Item / Model must not exceed 200 characters."}), 400



    allowed_salespeople = get_allowed_salespeople()
    if sales_person:
        if sales_person not in allowed_salespeople:
            return jsonify({"success": False, "message": "Invalid Sales Person selected."}), 400

    if payment_mode not in ['Cash', 'Card', 'E-Wallet']:
        return jsonify({"success": False, "message": "Invalid Payment Mode selected."}), 400

    # Total Bill Amount
    if total_bill_amount is None or str(total_bill_amount).strip() == "":
        return jsonify({"success": False, "message": "Total Bill Amount is required."}), 400
    try:
        total_val = float(total_bill_amount)
        if total_val <= 0:
            return jsonify({"success": False, "message": "Total Bill Amount must be greater than zero."}), 400
    except ValueError:
        return jsonify({"success": False, "message": "Total Bill Amount must be a valid number."}), 400

    # Amount Received
    if amount_received is None or str(amount_received).strip() == "":
        return jsonify({"success": False, "message": "Amount Received is required."}), 400
    try:
        received_val = float(amount_received)
        if received_val < 0:
            return jsonify({"success": False, "message": "Amount Received cannot be negative."}), 400
        if received_val > total_val:
            return jsonify({"success": False, "message": "Amount Received cannot be greater than Total Bill Amount."}), 400
    except ValueError:
        return jsonify({"success": False, "message": "Amount Received must be a valid number."}), 400

    pending_val = round(total_val - received_val, 2)
    
    # Pending validations
    if pending_val > 0:
        if pending_from not in ['Customer', 'Employee / Salesperson']:
            return jsonify({"success": False, "message": "Pending From is required and must be selected when there is a pending balance."}), 400
        if not due_date:
            return jsonify({"success": False, "message": "Due Date is required when there is a pending balance."}), 400
        
        if payment_status_override in ['Fully Paid', 'Partially Paid', 'Pending', 'Overdue']:
            payment_status = payment_status_override
        else:
            payment_status = "Partially Paid" if received_val > 0 else "Pending"
            
        if settlement_status_override in ['Settled', 'Unsettled']:
            settlement_status = settlement_status_override
        else:
            settlement_status = "Unsettled"
    else:
        pending_from = "No Pending"
        due_date = ""
        pending_person_name = ""
        
        if payment_status_override in ['Fully Paid', 'Partially Paid', 'Pending', 'Overdue']:
            payment_status = payment_status_override
        else:
            payment_status = "Fully Paid"
            
        if settlement_status_override in ['Settled', 'Unsettled']:
            settlement_status = settlement_status_override
        else:
            settlement_status = "Settled"

    if len(remarks) > 500:
        return jsonify({"success": False, "message": "Remarks must not exceed 500 characters."}), 400

    # Generate incremental PAY-XXXX ID
    payments = load_payments()
    next_num = 1001
    if payments:
        for p in payments:
            pid = p.get('payment_id', '')
            if pid.startswith('PAY-'):
                try:
                    num = int(pid.split('-')[1])
                    if num >= next_num:
                        next_num = num + 1
                except (ValueError, IndexError):
                    pass
    payment_id = f"PAY-{next_num}"

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Initial history log
    history_entry = {
        "date_time": now_str,
        "amount_added": received_val,
        "received_by": session['username'],
        "remarks": "Initial record creation."
    }

    new_payment = {
        "payment_id": payment_id,
        "customer_name": customer_name,
        "mobile_number": mobile_number,
        "invoice_number": invoice_number,
        "item_model": item_model,
        "imei_number": imei_number,
        "sales_person": sales_person,
        "payment_mode": payment_mode,
        "total_bill_amount": total_val,
        "amount_received": received_val,
        "pending_amount": pending_val,
        "pending_from": pending_from,
        "pending_person_name": pending_person_name,
        "due_date": due_date,
        "payment_status": payment_status,
        "settlement_status": settlement_status,
        "remarks": remarks,
        "created_at": now_str,
        "updated_at": now_str,
        "created_by": session['username'],
        "payment_history": [history_entry]
    }

    payments.append(new_payment)
    save_payments(payments)
    
    return jsonify({"success": True, "message": "Payment record added successfully!", "data": new_payment})

@app.route('/api/payments/<payment_id>', methods=['PUT'])
@login_required
@permission_required('payment_tracker_edit')
def api_update_payment(payment_id):
    payments = load_payments()
    payment = None
    for p in payments:
        if p.get('payment_id') == payment_id:
            payment = p
            break
            
    if not payment:
        return jsonify({"success": False, "message": "Payment record not found."}), 404
        
    data = request.get_json() or {}
    
    # Read/validate edits
    customer_name = data.get('customer_name', '').strip()
    mobile_number = data.get('mobile_number', '').strip()
    invoice_number = data.get('invoice_number', '').strip()
    item_model = data.get('item_model', '').strip()
    imei_number = data.get('imei_number', '').strip()
    sales_person = data.get('sales_person', '').strip()
    payment_mode = data.get('payment_mode', '').strip()
    total_bill_amount = data.get('total_bill_amount')
    amount_received = data.get('amount_received')
    pending_from = data.get('pending_from', '').strip()
    pending_person_name = data.get('pending_person_name', '').strip()
    due_date = data.get('due_date', '').strip()
    remarks = data.get('remarks', '').strip()
    payment_status_override = data.get('payment_status', '').strip()
    settlement_status_override = data.get('settlement_status', '').strip()

    # Validations (same as POST)
    if not customer_name:
        return jsonify({"success": False, "message": "Customer Name is required."}), 400
    if len(customer_name) > 100:
        return jsonify({"success": False, "message": "Customer Name must not exceed 100 characters."}), 400
    customer_name = ' '.join(w.capitalize() for w in customer_name.split())

    if not mobile_number:
        return jsonify({"success": False, "message": "Mobile Number is required."}), 400
    if len(mobile_number) != 10 or not mobile_number.isdigit():
        return jsonify({"success": False, "message": "Mobile Number must be exactly 10 digits."}), 400

    if item_model and len(item_model) > 200:
        return jsonify({"success": False, "message": "Item / Model must not exceed 200 characters."}), 400



    allowed_salespeople = get_allowed_salespeople()
    if sales_person:
        if sales_person not in allowed_salespeople:
            return jsonify({"success": False, "message": "Invalid Sales Person selected."}), 400

    if payment_mode not in ['Cash', 'Card', 'E-Wallet']:
        return jsonify({"success": False, "message": "Invalid Payment Mode selected."}), 400

    # Total Bill
    try:
        total_val = float(total_bill_amount)
        if total_val <= 0:
            return jsonify({"success": False, "message": "Total Bill Amount must be greater than zero."}), 400
    except (ValueError, TypeError):
        return jsonify({"success": False, "message": "Total Bill Amount must be a valid number."}), 400

    # Amount Received
    try:
        received_val = float(amount_received)
        if received_val < 0:
            return jsonify({"success": False, "message": "Amount Received cannot be negative."}), 400
        if received_val > total_val:
            return jsonify({"success": False, "message": "Amount Received cannot be greater than Total Bill Amount."}), 400
    except (ValueError, TypeError):
        return jsonify({"success": False, "message": "Amount Received must be a valid number."}), 400

    pending_val = round(total_val - received_val, 2)

    # Pending checks
    if pending_val > 0:
        if pending_from not in ['Customer', 'Employee / Salesperson']:
            return jsonify({"success": False, "message": "Pending From is required and must be selected when there is a pending balance."}), 400
        if not due_date:
            return jsonify({"success": False, "message": "Due Date is required when there is a pending balance."}), 400
        
        if payment_status_override in ['Fully Paid', 'Partially Paid', 'Pending', 'Overdue']:
            payment_status = payment_status_override
        else:
            payment_status = "Partially Paid" if received_val > 0 else "Pending"
            
        if settlement_status_override in ['Settled', 'Unsettled']:
            settlement_status = settlement_status_override
        else:
            settlement_status = "Unsettled"
    else:
        pending_from = "No Pending"
        due_date = ""
        pending_person_name = ""
        
        if payment_status_override in ['Fully Paid', 'Partially Paid', 'Pending', 'Overdue']:
            payment_status = payment_status_override
        else:
            payment_status = "Fully Paid"
            
        if settlement_status_override in ['Settled', 'Unsettled']:
            settlement_status = settlement_status_override
        else:
            settlement_status = "Settled"

    if len(remarks) > 500:
        return jsonify({"success": False, "message": "Remarks must not exceed 500 characters."}), 400

    # If amount received changed, we log it in the history
    prev_received = float(payment.get('amount_received') or 0)
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    if received_val != prev_received:
        diff = received_val - prev_received
        payment['payment_history'].append({
            "date_time": now_str,
            "amount_added": round(diff, 2),
            "received_by": session['username'],
            "remarks": f"Amount adjusted via edit (changed from Rupee {prev_received} to Rupee {received_val})."
        })

    # Update fields
    payment.update({
        "customer_name": customer_name,
        "mobile_number": mobile_number,
        "invoice_number": invoice_number,
        "item_model": item_model,
        "imei_number": imei_number,
        "sales_person": sales_person,
        "payment_mode": payment_mode,
        "total_bill_amount": total_val,
        "amount_received": received_val,
        "pending_amount": pending_val,
        "pending_from": pending_from,
        "pending_person_name": pending_person_name,
        "due_date": due_date,
        "payment_status": payment_status,
        "settlement_status": settlement_status,
        "remarks": remarks,
        "updated_at": now_str
    })

    save_payments(payments)
    return jsonify({"success": True, "message": "Payment record updated successfully!", "data": payment})

@app.route('/api/payments/<payment_id>', methods=['DELETE'])
@login_required
@permission_required('payment_tracker_delete')
def api_delete_payment(payment_id):
    payments = load_payments()
    updated_payments = [p for p in payments if p.get('payment_id') != payment_id]
    
    if len(updated_payments) == len(payments):
        return jsonify({"success": False, "message": "Payment record not found."}), 404
        
    save_payments(updated_payments)
    return jsonify({"success": True, "message": "Payment record deleted successfully!"})

@app.route('/api/payments/<payment_id>/partial-payment', methods=['POST'])
@login_required
@permission_required('payment_tracker_edit')
def api_partial_payment(payment_id):
    payments = load_payments()
    payment = None
    for p in payments:
        if p.get('payment_id') == payment_id:
            payment = p
            break
            
    if not payment:
        return jsonify({"success": False, "message": "Payment record not found."}), 404
        
    data = request.get_json() or {}
    amount_added = data.get('amount_added')
    partial_remarks = data.get('remarks', '').strip()
    
    if amount_added is None or str(amount_added).strip() == "":
        return jsonify({"success": False, "message": "Amount added is required."}), 400
        
    try:
        added_val = float(amount_added)
        if added_val <= 0:
            return jsonify({"success": False, "message": "Amount added must be greater than zero."}), 400
            
        pending_val = float(payment.get('pending_amount') or 0)
        if added_val > pending_val:
            return jsonify({"success": False, "message": f"Amount added cannot exceed the pending amount of Rupee {pending_val}."}), 400
    except ValueError:
        return jsonify({"success": False, "message": "Amount added must be a valid number."}), 400
        
    # Update amounts
    new_received = round(float(payment.get('amount_received') or 0) + added_val, 2)
    new_pending = round(float(payment.get('total_bill_amount') or 0) - new_received, 2)
    
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    history_entry = {
        "date_time": now_str,
        "amount_added": added_val,
        "received_by": session['username'],
        "remarks": partial_remarks or "Partial payment added."
    }
    
    payment['amount_received'] = new_received
    payment['pending_amount'] = new_pending
    payment['payment_history'].append(history_entry)
    payment['updated_at'] = now_str
    
    # Recalculate status
    if new_pending == 0:
        payment['payment_status'] = 'Fully Paid'
        payment['settlement_status'] = 'Settled'
        payment['pending_from'] = 'No Pending'
        payment['due_date'] = ''
        payment['pending_person_name'] = ''
    else:
        payment['payment_status'] = 'Partially Paid'
        
    save_payments(payments)
    return jsonify({"success": True, "message": "Partial payment added successfully!", "data": payment})

@app.route('/api/payments/<payment_id>/mark-paid', methods=['POST'])
@login_required
@permission_required('payment_tracker_edit')
def api_mark_paid(payment_id):
    payments = load_payments()
    payment = None
    for p in payments:
        if p.get('payment_id') == payment_id:
            payment = p
            break
            
    if not payment:
        return jsonify({"success": False, "message": "Payment record not found."}), 404
        
    pending_val = float(payment.get('pending_amount') or 0)
    if pending_val <= 0:
        return jsonify({"success": False, "message": "This payment is already fully paid."}), 400
        
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    history_entry = {
        "date_time": now_str,
        "amount_added": pending_val,
        "received_by": session['username'],
        "remarks": "Marked as fully paid (balance cleared)."
    }
    
    payment['amount_received'] = float(payment.get('total_bill_amount') or 0)
    payment['pending_amount'] = 0.0
    payment['payment_status'] = 'Fully Paid'
    payment['settlement_status'] = 'Settled'
    payment['pending_from'] = 'No Pending'
    payment['due_date'] = ''
    payment['pending_person_name'] = ''
    payment['payment_history'].append(history_entry)
    payment['updated_at'] = now_str
    
    save_payments(payments)
    return jsonify({"success": True, "message": "Payment marked as fully paid successfully!", "data": payment})

# --- USER MANAGEMENT ENDPOINTS ---

@app.route('/admin/users')
@login_required
@permission_required('settings_access')
def admin_users_view():
    return render_template('admin_users.html', username=session['username'], role=session['role'])

@app.route('/admin/database-editor')
@login_required
@permission_required('settings_access')
def database_editor_view():
    return render_template('database_editor.html', username=session['username'], role=session['role'])

def get_allowed_salespeople():
    users = load_users()
    allowed = []
    for u in users:
        if not isinstance(u, dict):
            continue
        if u.get('status', 'active') != 'active':
            continue
        name = (u.get('username') or '').strip()
        employee_id = (u.get('employee_id') or '').strip()
        if not name or not employee_id:
            continue
        
        # 1. Add employee_id based displays (what UI dropdown sends)
        allowed.append(f"{name} - {employee_id}")
        allowed.append(f"{name} \u2014 {employee_id}")
        allowed.append(f"{name} \u2013 {employee_id}")
        
        # 2. Add job_title based displays (for legacy/admin records)
        job_title = (u.get('job_title') or '').strip()
        if job_title:
            allowed.append(f"{name} - {job_title}")
            allowed.append(f"{name} \u2014 {job_title}")
            allowed.append(f"{name} \u2013 {job_title}")
    
    # Keep legacy hardcoded values for backward compatibility
    legacy = [
        "Bharath Kumar - Manager",
        "Ramesh T - Cashier",
        "Mohammad Farooq - Staff",
        "Farooq - Staff",
        "Abhishek - OPPO",
        "Azeem - VIVO",
        "Azeem - Vivo",
        "Rabiya - Xiaomi",
        "Finance Promoter"
    ]
    return list(set(allowed + legacy))

# Sales Persons / Employees list — accessible to any logged-in user, no admin permission required
@app.route('/api/sales-persons', methods=['GET'])
@login_required
def api_sales_persons():
    """Return active sales persons/employees from Supabase users table."""
    users = load_users()
    sales_persons = []
    for u in users:
        if u.get('status', 'active').lower() != 'active':
            continue
        username = u.get('username', '')
        employee_id = u.get('employee_id', '')
        display_name = f"{username} - {employee_id}"
        sales_persons.append({
            "user_id": u.get('user_id') or u.get('id', ''),
            "username": username,
            "employee_id": employee_id,
            "role": u.get('role', ''),
            "display_name": display_name
        })
    return jsonify(sales_persons)

@app.route('/api/staff-list', methods=['GET'])
@login_required
def api_staff_list():
    """Fallback legacy route for active staff list."""
    users = load_users()
    staff = []
    for u in users:
        if u.get('status', 'active').lower() != 'active':
            continue
        username = u.get('username', '')
        employee_id = u.get('employee_id', '')
        display = f"{username} - {employee_id}"
        staff.append({
            'username': username,
            'role': u.get('role', ''),
            'job_title': u.get('job_title', '') or 'Staff',
            'display': display
        })
    return jsonify({'success': True, 'data': staff})

@app.route('/api/users', methods=['GET'])
@login_required
@permission_required('settings_access')
def api_get_users():
    users = load_users()
    # Strip password hashes for safety
    cleaned_users = []
    for u in users:
        cleaned = u.copy()
        cleaned.pop('password_hash', None)
        cleaned_users.append(cleaned)
    return jsonify({"success": True, "data": cleaned_users})

@app.route('/api/users', methods=['POST'])
@login_required
@permission_required('user_create')
def api_create_user():
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    employee_id = (data.get('employee_id') or '').strip()
    password = (data.get('password') or '').strip()
    role = (data.get('role') or '').strip()
    job_title = (data.get('job_title') or '').strip() or None
    
    if not username or not employee_id or not password or not role:
        return jsonify({"success": False, "message": "All fields (Username, Employee ID, Password, Role) are required."}), 400
        
    if role not in ['super_admin', 'admin', 'store_employee']:
        return jsonify({"success": False, "message": "Invalid user role."}), 400
        
    users = load_users()
    if any(u.get('employee_id') == employee_id for u in users):
        return jsonify({"success": False, "message": f"Duplicate Employee ID: User with Employee ID {employee_id} already exists."}), 400
        
    # Generate unique ID USR-XXXX
    next_num = 1004
    for u in users:
        uid = u.get('user_id', '')
        if uid.startswith('USR-'):
            try:
                num = int(uid.split('-')[1])
                if num >= next_num:
                    next_num = num + 1
            except (ValueError, IndexError):
                pass
    user_id = f"USR-{next_num}"
    
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    new_user = {
        "user_id": user_id,
        "username": username,
        "employee_id": employee_id,
        "password_hash": generate_password_hash(password),
        "role": role,
        "job_title": job_title,
        "status": "active",
        "created_at": now_str,
        "updated_at": now_str
    }
    
    users.append(new_user)
    save_users(users)
    
    # Strip hash for return
    ret_data = new_user.copy()
    ret_data.pop('password_hash', None)
    return jsonify({"success": True, "message": "User created successfully!", "data": ret_data})

@app.route('/api/users/<user_id>', methods=['PUT'])
@login_required
@permission_required('user_update')
def api_update_user(user_id):
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    role = data.get('role', '').strip()
    status = data.get('status', '').strip()
    password = data.get('password', '').strip() # Optional password reset
    job_title = data.get('job_title')
    
    users = load_users()
    user = next((u for u in users if u.get('user_id') == user_id), None)
    
    if not user:
        return jsonify({"success": False, "message": "User not found."}), 404
        
    # Security check: current user cannot deactivate or downgrade themselves
    if user.get('employee_id') == session.get('employee_id'):
        if status == 'inactive':
            return jsonify({"success": False, "message": "You cannot deactivate your own account."}), 400
        if role != user.get('role'):
            return jsonify({"success": False, "message": "You cannot change your own role."}), 400
            
    if username:
        user['username'] = username
    if role:
        if role not in ['super_admin', 'admin', 'store_employee']:
            return jsonify({"success": False, "message": "Invalid user role."}), 400
        user['role'] = role
    if status:
        if status not in ['active', 'inactive']:
            return jsonify({"success": False, "message": "Invalid status."}), 400
        user['status'] = status
    if password:
        user['password_hash'] = generate_password_hash(password)
    if job_title is not None:
        user['job_title'] = job_title.strip() or None
        
    user['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    save_users(users)
    
    ret_data = user.copy()
    ret_data.pop('password_hash', None)
    return jsonify({"success": True, "message": "User updated successfully!", "data": ret_data})

@app.route('/api/users/<user_id>', methods=['DELETE'])
@login_required
@permission_required('user_delete')
def api_delete_user(user_id):
    users = load_users()
    user = next((u for u in users if u.get('user_id') == user_id), None)
    
    if not user:
        return jsonify({"success": False, "message": "User not found."}), 404
        
    if user.get('employee_id') == session.get('employee_id'):
        return jsonify({"success": False, "message": "You cannot delete your own account."}), 400
        
    updated_users = [u for u in users if u.get('user_id') != user_id]
    save_users(updated_users)
    return jsonify({"success": True, "message": "User deleted successfully!"})

def load_edit_requests():
    """Load edit requests list from Supabase."""
    return db_load_edit_requests()

def save_edit_requests(requests_list):
    """Save edit requests list to Supabase."""
    db_save_edit_requests(requests_list)


@app.route('/admin/edit-requests')
@login_required
@permission_required('edit_request_view')
def admin_edit_requests_view():
    return render_template('admin_edit_requests.html', username=session['username'], role=session['role'])

@app.route('/api/edit-requests', methods=['GET'])
@login_required
@permission_required('edit_request_view')
def api_get_edit_requests():
    requests_list = load_edit_requests()
    # Sort requests: newer first
    requests_list.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return jsonify({"success": True, "data": requests_list})

@app.route('/api/customers/<customer_id>/edit-request', methods=['POST'])
@login_required
@permission_required('customer_edit_request_create')
def api_create_edit_request(customer_id):
    data = request.get_json() or {}
    reason = data.get('reason', '').strip()
    proposed_data = data.get('proposed_data', {})
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if not reason:
        return jsonify({"success": False, "message": "Reason for edit request is required."}), 400

    if not proposed_data:
        return jsonify({"success": False, "message": "Proposed changes are required."}), 400

    customers = load_customers()
    customer = next((c for c in customers if c.get('customer_id') == customer_id), None)
    if not customer:
        return jsonify({"success": False, "message": "Customer record not found."}), 404

    if customer.get('record_locked'):
        return jsonify({"success": False, "message": "This customer record is locked (billing verified) and cannot accept edit requests."}), 403

    # Check for duplicate pending requests by the same user for this customer
    requests_list = load_edit_requests()
    duplicate = next((r for r in requests_list if r.get('customer_id') == customer_id and r.get('requested_by') == session.get('employee_id') and r.get('status') == 'pending'), None)
    if duplicate:
        return jsonify({"success": False, "message": "You already have a pending edit request for this customer."}), 400

    # Validate proposed data strictly on backend
    p_name = proposed_data.get('customer_name', '').strip()
    p_mobile = proposed_data.get('mobile_number', '').strip()
    p_t_mode = proposed_data.get('transaction_mode', '').strip()
    p_item = proposed_data.get('item_model', '').strip()
    p_imei = proposed_data.get('imei_number', '').strip()
    p_sales = proposed_data.get('sales_person', '').strip()

    if not p_name or not p_mobile or not p_item or not p_imei or not p_t_mode or not p_sales:
        return jsonify({"success": False, "message": "Name, Mobile, Model, IMEI, Transaction Mode, and Sales Person are required in proposed changes."}), 400

    if len(p_mobile) != 10 or not p_mobile.isdigit():
        return jsonify({"success": False, "message": "Proposed Mobile Number must be exactly 10 digits."}), 400



    p_name = ' '.join(word.capitalize() for word in p_name.split())
    proposed_data['customer_name'] = p_name

    if p_t_mode == 'Finance':
        p_provider = proposed_data.get('finance_provider', '').strip()
        p_dp_mode = proposed_data.get('down_payment_mode', '').strip()
        p_dp_val_raw = proposed_data.get('down_payment_value', 0)
        
        if not p_provider or p_dp_mode not in ['Cash', 'Card', 'E-Wallet']:
            return jsonify({"success": False, "message": "Proposed Finance Provider and Down Payment Mode are required."}), 400
        try:
            p_dp_val = float(p_dp_val_raw)
            if p_dp_val < 0: raise ValueError
            proposed_data['down_payment_value'] = p_dp_val
        except ValueError:
            return jsonify({"success": False, "message": "Proposed Down Payment Value must be a valid positive number."}), 400
        
        proposed_data['cash_amount'] = 0.0
        proposed_data['card_amount'] = 0.0
        proposed_data['ewallet_amount'] = 0.0
        proposed_data['total_amount_received'] = p_dp_val
    else:
        proposed_data['finance_provider'] = ""
        proposed_data['down_payment_mode'] = ""
        proposed_data['down_payment_value'] = 0.0
        
        try:
            p_cash = float(proposed_data.get('cash_amount', 0) or 0)
            p_card = float(proposed_data.get('card_amount', 0) or 0)
            p_ewallet = float(proposed_data.get('ewallet_amount', 0) or 0)
            if p_cash < 0 or p_card < 0 or p_ewallet < 0: raise ValueError
            proposed_data['cash_amount'] = p_cash
            proposed_data['card_amount'] = p_card
            proposed_data['ewallet_amount'] = p_ewallet
        except ValueError:
            return jsonify({"success": False, "message": "Proposed payment split amounts must be positive numbers."}), 400
            
        proposed_data['total_amount_received'] = round(p_cash + p_card + p_ewallet, 2)
        if proposed_data['total_amount_received'] <= 0:
            return jsonify({"success": False, "message": "Proposed payment split must have at least one amount greater than zero."}), 400

    p_exch_status = proposed_data.get('exchange_status', '').strip()
    if p_exch_status not in ['Yes', 'No']:
        return jsonify({"success": False, "message": "Exchange Status must be Yes or No."}), 400
    if p_exch_status == 'Yes':
        p_exch_brand = proposed_data.get('exchange_brand', '').strip()
        if not p_exch_brand:
            return jsonify({"success": False, "message": "Proposed Exchange Brand is required when Exchange is Yes."}), 400
        try:
            p_exch_val = float(proposed_data.get('exchange_value', 0) or 0)
            if p_exch_val < 0: raise ValueError
            proposed_data['exchange_value'] = p_exch_val
        except ValueError:
            return jsonify({"success": False, "message": "Proposed Exchange Value must be a valid positive number."}), 400
    else:
        proposed_data['exchange_brand'] = ""
        proposed_data['exchange_value'] = 0.0

    proposed_data['customer_id'] = customer_id
    proposed_data['created_by'] = customer.get('created_by', 'admin')
    proposed_data['created_at'] = customer.get('created_at', now_str)

    next_num = 1001
    if requests_list:
        for r in requests_list:
            rid = r.get('request_id', '')
            if rid.startswith('REQ-'):
                try:
                    num = int(rid.split('-')[1])
                    if num >= next_num:
                        next_num = num + 1
                except (ValueError, IndexError):
                    pass
    request_id = f"REQ-{next_num}"

    new_request = {
        "request_id": request_id,
        "customer_id": customer_id,
        "requested_by": session.get('employee_id', 'admin'),
        "requested_role": session.get('role', 'store_employee'),
        "original_data": customer,
        "proposed_data": proposed_data,
        "reason": reason,
        "status": "pending",
        "admin_remarks": "",
        "approved_by": "",
        "approved_at": "",
        "rejected_by": "",
        "rejected_at": "",
        "created_at": now_str
    }

    requests_list.append(new_request)
    save_edit_requests(requests_list)

    return jsonify({"success": True, "message": "Edit request submitted successfully. Waiting for admin approval.", "data": new_request})

@app.route('/api/edit-requests/<request_id>', methods=['DELETE'])
@login_required
@permission_required('edit_request_reject')
def api_delete_edit_request(request_id):
    requests_list = load_edit_requests()
    req = next((r for r in requests_list if r.get('request_id') == request_id), None)
    
    if not req:
        return jsonify({"success": False, "message": "Edit request not found."}), 404
        
    updated_list = [r for r in requests_list if r.get('request_id') != request_id]
    save_edit_requests(updated_list)
    return jsonify({"success": True, "message": "Edit request deleted successfully!"})

@app.route('/api/edit-requests/clear-history', methods=['POST'])
@login_required
@permission_required('edit_request_reject')
def api_clear_edit_requests_history():
    requests_list = load_edit_requests()
    pending_requests = [r for r in requests_list if r.get('status') == 'pending']
    save_edit_requests(pending_requests)
    return jsonify({"success": True, "message": "All edit requests history cleared successfully!"})

@app.route('/api/edit-requests/<request_id>/approve', methods=['POST'])
@login_required
@permission_required('edit_request_approve')
def api_approve_edit_request(request_id):
    data = request.get_json(silent=True) or {}
    admin_remarks = data.get('admin_remarks', '').strip()

    requests_list = load_edit_requests()
    req = next((r for r in requests_list if r.get('request_id') == request_id), None)
    
    if not req:
        return jsonify({"success": False, "message": "Edit request not found."}), 404
        
    if req.get('status') != 'pending':
        return jsonify({"success": False, "message": f"This request is already {req.get('status')}."}), 400

    customers = load_customers()
    customer_index = next((i for i, c in enumerate(customers) if c.get('customer_id') == req.get('customer_id')), -1)
    
    if customer_index == -1:
        return jsonify({"success": False, "message": "Original customer record not found."}), 404

    orig_cust = customers[customer_index]
    if orig_cust.get('record_locked'):
        return jsonify({"success": False, "message": "The original customer record is locked (billing verified) and cannot be updated."}), 403

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    proposed = req.get('proposed_data', {})
    
    # Overwrite original customer record, but preserve customer_id, created_at, created_by
    proposed['customer_id'] = orig_cust.get('customer_id')
    proposed['created_at'] = orig_cust.get('created_at')
    proposed['created_by'] = orig_cust.get('created_by')
    proposed['updated_at'] = now_str
    
    # Preserve billing verification and locking fields
    for field in [
        'billing_status', 'billing_verified', 'billing_verified_by', 'billing_verified_at',
        'billing_admin_remarks', 'record_locked', 'locked_by', 'locked_at',
        'reopened_by', 'reopened_at', 'reopen_reason'
    ]:
        if field in orig_cust:
            proposed[field] = orig_cust[field]
            
    customers[customer_index] = proposed
    save_customers(customers)
    
    # Update request entry
    req['status'] = 'approved'
    req['approved_by'] = session.get('employee_id', 'admin')
    req['approved_at'] = now_str
    if admin_remarks:
        req['admin_remarks'] = admin_remarks
    save_edit_requests(requests_list)
    
    return jsonify({"success": True, "message": "Edit request approved and customer record updated successfully!", "customer": proposed})

@app.route('/api/edit-requests/<request_id>/reject', methods=['POST'])
@login_required
@permission_required('edit_request_reject')
def api_reject_edit_request(request_id):
    data = request.get_json(silent=True) or {}
    admin_remarks = data.get('admin_remarks', '').strip()
    
    if not admin_remarks:
        return jsonify({"success": False, "message": "Admin remarks are required when rejecting."}), 400
        
    requests_list = load_edit_requests()
    req = next((r for r in requests_list if r.get('request_id') == request_id), None)
    
    if not req:
        return jsonify({"success": False, "message": "Edit request not found."}), 404
        
    if req.get('status') != 'pending':
        return jsonify({"success": False, "message": f"This request is already {req.get('status')}."}), 400

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Update request entry
    req['status'] = 'rejected'
    req['admin_remarks'] = admin_remarks
    req['rejected_by'] = session.get('employee_id', 'admin')
    req['rejected_at'] = now_str
    save_edit_requests(requests_list)
    
    return jsonify({"success": True, "message": "Edit request rejected successfully.", "data": req})

def migrate_customers_db():
    """Migrate customers database to the new Customer Entry Management System schema."""
    if not os.path.exists(CUSTOMERS_FILE):
        return
        
    try:
        with open(CUSTOMERS_FILE, 'r') as f:
            data = json.load(f)
            
        if not isinstance(data, list):
            return
            
        modified = False
        migrated_list = []
        for index, c in enumerate(data):
            c_modified = False
            
            # Check legacy primary key mapping
            legacy_id = c.get('id', index + 1)
            cust_id = c.get('customer_id')
            if not cust_id:
                cust_id = f"CUST-{1000 + legacy_id}"
                c_modified = True

            # Normalize transaction mode
            t_mode = c.get('transaction_mode', c.get('transactionMode', ''))
            if t_mode in ['EMI', 'Finance']:
                normal_t_mode = 'Finance'
            else:
                normal_t_mode = 'Non-Finance'
            if normal_t_mode != c.get('transaction_mode'):
                c_modified = True

            # Parse amounts safely
            legacy_amt_rec = c.get('amount_received', c.get('amountReceived', '0'))
            try:
                amt_received_val = float(legacy_amt_rec)
            except ValueError:
                amt_received_val = 0.0

            legacy_dp = c.get('down_payment', c.get('downPayment', '0'))
            try:
                dp_val = float(legacy_dp)
            except ValueError:
                dp_val = 0.0

            # Default payment splits based on payment mode
            pay_mode = c.get('payment_mode', c.get('paymentMode', 'Cash'))
            cash_amt = float(c.get('cash_amount', amt_received_val if pay_mode == 'Cash' else 0.0))
            card_amt = float(c.get('card_amount', amt_received_val if pay_mode == 'Card' else 0.0))
            ewallet_amt = float(c.get('ewallet_amount', amt_received_val if pay_mode == 'E-Wallet' else 0.0))
            
            if 'cash_amount' not in c or 'card_amount' not in c or 'ewallet_amount' not in c:
                c_modified = True

            total_rec = cash_amt + card_amt + ewallet_amt
            if normal_t_mode == 'Finance':
                total_rec = dp_val
            
            if 'total_amount_received' not in c:
                c_modified = True

            # Exchange brand and value
            exch_status = c.get('exchange_status', c.get('exchangeStatus', 'No'))
            exch_brand = c.get('exchange_brand', '')
            exch_val = float(c.get('exchange_value', 0.0))
            if 'exchange_brand' not in c or 'exchange_value' not in c:
                c_modified = True

            # Down payment mode
            dp_mode = c.get('down_payment_mode', pay_mode if normal_t_mode == 'Finance' else '')
            if 'down_payment_mode' not in c:
                c_modified = True

            # Auditing timestamps
            c_at = c.get('created_at')
            if not c_at:
                date_val = c.get('date', datetime.now().strftime('%Y-%m-%d'))
                time_val = c.get('time', datetime.now().strftime('%H:%M:%S'))
                c_at = f"{date_val} {time_val}"
                c_modified = True

            u_at = c.get('updated_at', c_at)
            c_by = c.get('created_by', 'admin')

            if c_modified:
                modified = True

            migrated_customer = {
                "customer_id": cust_id,
                "customer_name": ' '.join(w.capitalize() for w in c.get('customer_name', c.get('customerName', 'Unknown')).split()),
                "mobile_number": c.get('mobile_number', c.get('mobileNumber', '')),
                "item_model": c.get('item_model', c.get('itemModel', '')),
                "imei_number": c.get('imei_number', c.get('imei', '')),
                "transaction_mode": normal_t_mode,
                "finance_provider": c.get('finance_provider', c.get('financeOption', '')),
                "down_payment_mode": dp_mode,
                "down_payment_value": dp_val,
                "cash_amount": cash_amt,
                "card_amount": card_amt,
                "ewallet_amount": ewallet_amt,
                "total_amount_received": total_rec,
                "exchange_status": exch_status,
                "exchange_brand": exch_brand,
                "exchange_value": exch_val,
                "sales_person": c.get('sales_person', c.get('salesPerson', '')),
                "remarks": c.get('remarks', ''),
                "created_by": c_by,
                "created_at": c_at,
                "updated_at": u_at
            }
            migrated_list.append(migrated_customer)
                
        if modified:
            save_customers(migrated_list)
            print("Database migrated successfully to new CEMS JSON schema.")
    except Exception as e:
        print(f"Error migrating database: {e}")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
